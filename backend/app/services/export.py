from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.models import BudgetCategory, BudgetItem, BudgetSubCategory, Expense, Project
from app.services.budget import get_monthly_stats, get_project_summary


def generate_expense_excel(db: Session, project_id: int) -> bytes:
    project: Project = db.get(Project, project_id)
    summary = get_project_summary(db, project_id)

    wb = Workbook()

    header_fill = PatternFill(start_color="F97316", end_color="F97316", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    subtotal_fill = PatternFill(start_color="FED7AA", end_color="FED7AA", fill_type="solid")
    total_fill = PatternFill(start_color="EA580C", end_color="EA580C", fill_type="solid")
    total_font = Font(bold=True, color="FFFFFF")
    even_fill = PatternFill(start_color="FFF7ED", end_color="FFF7ED", fill_type="solid")
    sub_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

    # ── 시트1: 지출내역 ────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "지출내역"

    headers = ["번호", "날짜", "세목", "세세목", "품목", "내용(적요)", "지출처", "금액"]
    col_widths = [6, 14, 16, 16, 16, 28, 18, 16]

    for col, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws1.column_dimensions[get_column_letter(col)].width = width

    expenses = (
        db.query(Expense)
        .filter(Expense.project_id == project_id)
        .order_by(Expense.expense_date)
        .all()
    )

    row = 2
    grand_total = 0

    for seq, expense in enumerate(expenses, start=1):
        item = expense.budget_item
        sub = item.sub_category if item else None
        cat = sub.category if sub else None
        is_even = seq % 2 == 0
        values = [
            seq,
            expense.expense_date.strftime("%Y-%m-%d") if isinstance(expense.expense_date, date) else str(expense.expense_date),
            cat.name if cat else "",
            sub.name if sub else "",
            item.name if item else "",
            expense.description,
            expense.vendor or "",
            expense.amount,
        ]
        for col, val in enumerate(values, start=1):
            cell = ws1.cell(row=row, column=col, value=val)
            if is_even:
                cell.fill = even_fill
            if col == 8:
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")
        grand_total += expense.amount
        row += 1

    # 합계 행
    total_cell = ws1.cell(row=row, column=1, value="합  계")
    total_cell.font = total_font
    total_cell.fill = total_fill
    for col in range(2, 8):
        ws1.cell(row=row, column=col).fill = total_fill
    amount_cell = ws1.cell(row=row, column=8, value=grand_total)
    amount_cell.number_format = "#,##0"
    amount_cell.alignment = Alignment(horizontal="right")
    amount_cell.font = total_font
    amount_cell.fill = total_fill

    # ── 시트2: 집행현황 요약 (세목별) ─────────────────────────
    ws2 = wb.create_sheet("집행현황 요약")

    ws2.cell(row=1, column=1, value=f"사업명: {project.name}").font = Font(bold=True, size=13)
    ws2.cell(row=2, column=1, value=f"총 예산: {project.total_budget:,}원").font = Font(bold=True)

    headers2 = ["세목", "계획금액", "집행액", "잔액", "집행률(%)"]
    col_widths2 = [22, 18, 18, 18, 14]
    for col, (header, width) in enumerate(zip(headers2, col_widths2), start=1):
        cell = ws2.cell(row=4, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws2.column_dimensions[get_column_letter(col)].width = width

    for i, cat in enumerate(summary.categories, start=5):
        ws2.cell(row=i, column=1, value=cat.category_name)
        ws2.cell(row=i, column=2, value=cat.allocated_amount).number_format = "#,##0"
        ws2.cell(row=i, column=3, value=cat.spent_amount).number_format = "#,##0"
        ws2.cell(row=i, column=4, value=cat.remaining).number_format = "#,##0"
        rate_cell = ws2.cell(row=i, column=5, value=round(cat.execution_rate, 1))
        rate_cell.number_format = "0.0"
        rate_cell.alignment = Alignment(horizontal="right")
        if cat.execution_rate >= 90:
            for col in range(1, 6):
                ws2.cell(row=i, column=col).fill = PatternFill(start_color="FECACA", end_color="FECACA", fill_type="solid")

    summary_row = 5 + len(summary.categories)
    ws2.cell(row=summary_row, column=1, value="합  계").font = total_font
    ws2.cell(row=summary_row, column=2, value=sum(c.allocated_amount for c in summary.categories)).number_format = "#,##0"
    ws2.cell(row=summary_row, column=3, value=summary.total_spent).number_format = "#,##0"
    ws2.cell(row=summary_row, column=4, value=summary.total_remaining).number_format = "#,##0"
    ws2.cell(row=summary_row, column=5, value=round(summary.execution_rate, 1)).number_format = "0.0"
    for col in range(1, 6):
        ws2.cell(row=summary_row, column=col).fill = total_fill
        ws2.cell(row=summary_row, column=col).font = total_font

    # ── 시트3: 예산계획 산출내역 (세목→세세목→품목) ──────────
    ws3 = wb.create_sheet("예산계획 산출내역")

    ws3.cell(row=1, column=1, value=f"사업명: {project.name}").font = Font(bold=True, size=13)
    ws3.cell(row=2, column=1, value="산출내역 (단가 × 수량 = 계획금액)")

    plan_headers = ["세목", "세세목", "품목", "단가", "수량", "계획금액", "비고"]
    plan_widths = [18, 18, 18, 14, 10, 16, 16]
    for col, (header, width) in enumerate(zip(plan_headers, plan_widths), start=1):
        cell = ws3.cell(row=4, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws3.column_dimensions[get_column_letter(col)].width = width

    categories = (
        db.query(BudgetCategory)
        .filter_by(project_id=project_id)
        .order_by(BudgetCategory.order_index)
        .all()
    )

    plan_row = 5
    plan_grand = 0

    for cat in categories:
        sub_categories = (
            db.query(BudgetSubCategory)
            .filter_by(category_id=cat.id)
            .order_by(BudgetSubCategory.order_index)
            .all()
        )
        cat_total = 0
        cat_start_row = plan_row

        for sub in sub_categories:
            items = db.query(BudgetItem).filter_by(sub_category_id=sub.id).all()
            sub_total = 0
            sub_start_row = plan_row

            for item in items:
                planned = item.unit_price * item.quantity
                ws3.cell(row=plan_row, column=3, value=item.name)
                cell_price = ws3.cell(row=plan_row, column=4, value=item.unit_price)
                cell_price.number_format = "#,##0"
                ws3.cell(row=plan_row, column=5, value=item.quantity)
                cell_planned = ws3.cell(row=plan_row, column=6, value=planned)
                cell_planned.number_format = "#,##0"
                cell_planned.alignment = Alignment(horizontal="right")
                ws3.cell(row=plan_row, column=7, value=item.note or "")
                sub_total += planned
                plan_row += 1

            # 세세목 소계
            sub_cell = ws3.cell(row=sub_start_row, column=2, value=sub.name)
            sub_cell.font = Font(bold=True)
            if len(items) > 1:
                ws3.merge_cells(f"B{sub_start_row}:B{plan_row - 1}")
                sub_cell.alignment = Alignment(vertical="center")
            sub_total_cell = ws3.cell(row=sub_start_row, column=6)
            if not items:
                sub_total_cell.value = 0
                sub_total_cell.number_format = "#,##0"
            for r in range(sub_start_row, plan_row):
                for c in [1, 2]:
                    ws3.cell(row=r, column=c).fill = sub_fill
            cat_total += sub_total

        # 세목 병합
        if plan_row > cat_start_row:
            cat_cell = ws3.cell(row=cat_start_row, column=1, value=cat.name)
            cat_cell.font = Font(bold=True)
            if plan_row - cat_start_row > 1:
                ws3.merge_cells(f"A{cat_start_row}:A{plan_row - 1}")
                cat_cell.alignment = Alignment(vertical="center")
            for r in range(cat_start_row, plan_row):
                ws3.cell(row=r, column=1).fill = PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid")

        plan_grand += cat_total

    # 전체 합계
    ws3.cell(row=plan_row, column=1, value="합  계").font = total_font
    ws3.cell(row=plan_row, column=1).fill = total_fill
    total_plan_cell = ws3.cell(row=plan_row, column=6, value=plan_grand)
    total_plan_cell.number_format = "#,##0"
    total_plan_cell.alignment = Alignment(horizontal="right")
    total_plan_cell.font = total_font
    total_plan_cell.fill = total_fill
    for col in [2, 3, 4, 5, 7]:
        ws3.cell(row=plan_row, column=col).fill = total_fill

    # ── 시트4: 월별 집행현황 ──────────────────────────────────
    ws4 = wb.create_sheet("월별 집행현황")

    monthly_stats = get_monthly_stats(db, project_id)
    months = sorted({s.month for s in monthly_stats})
    categories_in_monthly = sorted({s.category_name for s in monthly_stats})
    stat_map = {(s.month, s.category_name): s.total_amount for s in monthly_stats}

    ws4.cell(row=1, column=1, value=f"사업명: {project.name}").font = Font(bold=True, size=13)
    ws4.cell(row=2, column=1, value="월별 집행현황 (단위: 원)").font = Font(bold=True)

    header_row = 4
    ws4.cell(row=header_row, column=1, value="세목").fill = header_fill
    ws4.cell(row=header_row, column=1).font = header_font
    ws4.cell(row=header_row, column=1).alignment = Alignment(horizontal="center")
    ws4.column_dimensions["A"].width = 20

    for col_idx, month in enumerate(months, start=2):
        cell = ws4.cell(row=header_row, column=col_idx, value=month)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws4.column_dimensions[get_column_letter(col_idx)].width = 14

    total_col = len(months) + 2
    total_header = ws4.cell(row=header_row, column=total_col, value="합계")
    total_header.fill = header_fill
    total_header.font = header_font
    total_header.alignment = Alignment(horizontal="center")
    ws4.column_dimensions[get_column_letter(total_col)].width = 14

    for row_idx, cat_name in enumerate(categories_in_monthly, start=header_row + 1):
        ws4.cell(row=row_idx, column=1, value=cat_name)
        row_total = 0
        for col_idx, month in enumerate(months, start=2):
            amount = stat_map.get((month, cat_name), 0)
            cell = ws4.cell(row=row_idx, column=col_idx, value=amount)
            cell.number_format = "#,##0"
            cell.alignment = Alignment(horizontal="right")
            row_total += amount
        total_cell = ws4.cell(row=row_idx, column=total_col, value=row_total)
        total_cell.number_format = "#,##0"
        total_cell.alignment = Alignment(horizontal="right")
        total_cell.font = Font(bold=True)

    grand_row = header_row + len(categories_in_monthly) + 1
    ws4.cell(row=grand_row, column=1, value="합  계").font = total_font
    ws4.cell(row=grand_row, column=1).fill = total_fill
    grand = 0
    for col_idx, month in enumerate(months, start=2):
        month_total = sum(stat_map.get((month, c), 0) for c in categories_in_monthly)
        cell = ws4.cell(row=grand_row, column=col_idx, value=month_total)
        cell.number_format = "#,##0"
        cell.alignment = Alignment(horizontal="right")
        cell.fill = total_fill
        cell.font = total_font
        grand += month_total
    grand_cell = ws4.cell(row=grand_row, column=total_col, value=grand)
    grand_cell.number_format = "#,##0"
    grand_cell.alignment = Alignment(horizontal="right")
    grand_cell.fill = total_fill
    grand_cell.font = total_font

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
